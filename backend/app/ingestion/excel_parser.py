"""XLSX parser for purchase/sales invoices.

Reads the first worksheet with openpyxl in read-only mode (cheap on
memory), converts each row to a dict, then reuses the CSV path's
per-row validation via a small in-memory CSV round-trip.

Same header contract as the CSV parser — see ``csv_parser.py``.
"""
from __future__ import annotations

import csv
import io
from typing import BinaryIO

from openpyxl import load_workbook

from app.ingestion.csv_parser import ParseResult, SchemaError, parse_invoice_csv


def parse_invoice_xlsx(
    fh: BinaryIO,
    gstin_profile_id: str,
    direction: str,
) -> ParseResult:
    """Parse an XLSX file. Only the first sheet is read."""
    wb = load_workbook(fh, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise SchemaError("XLSX has no active sheet")
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise SchemaError("XLSX is empty")

        header_str = [
            "" if h is None else str(h).strip() for h in header
        ]

        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(header_str)
        for row in rows:
            cleaned = [
                "" if v is None else (
                    v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)
                )
                for v in row
            ]
            # Skip trailing all-empty rows (Excel often leaves them).
            if not any(c for c in cleaned):
                continue
            w.writerow(cleaned)

        buf.seek(0)
        return parse_invoice_csv(
            buf,
            gstin_profile_id=gstin_profile_id,
            direction=direction,
        )
    finally:
        wb.close()
